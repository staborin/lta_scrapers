import logging
import os
import re
import requests
import csv
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


BASE_URL_TEMPLATE = (
    "https://competitions.lta.org.uk/ranking/category.aspx"
    "?id={rank_id}&category={category}&C{category}RFPC=&p={p}&ps=100"
)


MAX_PAGES = 15

# Where to dump debugging artifacts (html + png) on failure
ARTIFACT_DIR = Path("lta_artifacts")

WEBAPP_URL = os.environ.get("LTA_SHEETS_WEBAPP_URL", "").strip()
WEBAPP_SECRET = os.environ.get("LTA_SHEETS_WEBAPP_SECRET", "").strip()


def resolve_recent_form_rank_id(logger):
    """
    Resolve weekly rank_id for U9/U10 via:
    ranking → LTA Recent Form → 10U Boys
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        logger.info("Resolving U9/U10 rank_id via LTA Recent Form...")

        page.goto("https://competitions.lta.org.uk/ranking/", wait_until="domcontentloaded", timeout=60000)
        accept_cookies_if_needed(page, logger)

        # Click "LTA Recent Form"
        try:
            page.locator("a[href*='ranking.aspx?rid=303']").first.click()
        except Exception:
            page.locator("a:has-text('LTA Recent Form')").first.click()

        page.wait_for_load_state("domcontentloaded")
        accept_cookies_if_needed(page, logger)

        # Click "10U Boys"
        page.locator("a:has-text('10U Boys')").first.click()
        page.wait_for_load_state("domcontentloaded")

        final_url = page.url
        logger.info(f"[U9/U10] Resolved URL: {final_url}")

        qs = parse_qs(urlparse(final_url).query)
        if "id" not in qs or not qs["id"]:
            raise RuntimeError(f"Could not extract rank_id from URL: {final_url}")

        rank_id = int(qs["id"][0])
        logger.info(f"[U9/U10] Resolved rank_id={rank_id}")

        context.close()
        browser.close()

        return rank_id



def get_ranking_week(page, logger) -> str:
    """
    Returns the currently selected Ranking week, e.g. '6-2026'.
    Best source is the hidden <select name="...dlPublication...">.
    """
    sel = "select[name*='dlPublication']"
    try:
        if page.locator(sel).count():
            txt = page.eval_on_selector(
                sel,
                "s => (s.options[s.selectedIndex]?.textContent || '').trim()"
            )
            if txt:
                logger.info(f"Ranking week detected: {txt}")
                return txt
    except Exception:
        pass

    # Fallback: chosen widget (visible)
    try:
        fallback = page.locator("p:has(strong:has-text('Ranking week')) a.chosen-single span").first
        if fallback.count():
            txt = fallback.inner_text().strip()
            if txt:
                logger.info(f"Ranking week detected (fallback): {txt}")
                return txt
    except Exception:
        pass

    logger.warning("Could not detect Ranking week.")
    return ""




def write_rows_to_excel(header, rows, sheet_tab, logger, filename="LTA_MASTER_LIVE.xlsx"):
    excel_path = Path(__file__).resolve().parent / filename

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

    # Replace sheet if it already exists
    if sheet_tab in wb.sheetnames:
        ws = wb[sheet_tab]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_tab)

    ws.append(header)
    for r in rows:
        ws.append(r)

    wb.save(excel_path)
    logger.info(f"Wrote Excel tab '{sheet_tab}' in {excel_path}")



def find_rankings_table_selector(page):
    """
    Find the table that contains a 'Member ID' header.
    Returns a CSS selector like 'table:nth-of-type(2)' or None.
    """
    idx = page.evaluate(r"""
        () => {
          const tables = Array.from(document.querySelectorAll("table"));
          for (let i = 0; i < tables.length; i++) {
            const ths = Array.from(tables[i].querySelectorAll("thead th"))
              .map(th => (th.innerText || "").trim());
            if (ths.some(t => t.toLowerCase() === "member id")) {
              return i;  // zero-based index
            }
          }
          return null;
        }
    """)
    if idx is None:
        return None
    # CSS nth-of-type is 1-based
    return f"table:nth-of-type({idx + 1})"



def write_rows_to_sheet_via_webapp(header, rows, sheet_tab, clear_first, logger: logging.Logger):
    if not WEBAPP_URL:
        raise RuntimeError("Missing env var LTA_SHEETS_WEBAPP_URL")

    url = WEBAPP_URL
    if WEBAPP_SECRET:
        joiner = "&" if "?" in url else "?"
        url = f"{url}{joiner}secret={WEBAPP_SECRET}"

    payload = {
        "sheet": sheet_tab,
        "clearFirst": bool(clear_first),
        # Apps Script expects a rectangular 2D array including header row
        "rows": [header] + rows,
    }

    logger.info(f"Posting {len(rows)} rows to Google Sheet tab '{sheet_tab}'")
    r = requests.post(url, json=payload, timeout=120)
    r.raise_for_status()
    logger.info(f"Google Sheets response: {r.text.strip()}")


def setup_logger() -> logging.Logger:
    logger = logging.getLogger("lta_scrape_u9_u10")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        ch.setFormatter(fmt)
        logger.addHandler(ch)
    return logger


def clean_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def save_debug_artifacts(page, page_num: int, logger: logging.Logger, tag: str) -> None:
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = ARTIFACT_DIR / f"{ts}_{tag}_page_{page_num}_fail.html"
    png_path = ARTIFACT_DIR / f"{ts}_{tag}_page_{page_num}_fail.png"

    try:
        html_path.write_text(page.content(), encoding="utf-8")
        logger.info(f"Saved HTML: {html_path}")
    except Exception as e:
        logger.info(f"Failed to save HTML: {e}")

    try:
        page.screenshot(path=str(png_path), full_page=True)
        logger.info(f"Saved screenshot: {png_path}")
    except Exception as e:
        logger.info(f"Failed to save screenshot: {e}")


def accept_cookies_if_needed(page, logger):
    # Keep this pattern simple + defensive. If no banner, do nothing.
    try:
        # Common patterns on LTA pages (may vary)
        btn = page.locator("button:has-text('Accept')").first
        if btn.count() > 0:
            btn.click(timeout=2000)
            logger.info("Accepted cookies.")
    except Exception:
        pass


def get_headers_u9_u10(page):
    """
    U9/U10 tables don't always have a <thead>. If missing, the first tbody row acts as header.
    """
    # Try thead first (if present)
    headers = page.eval_on_selector_all(
        "table.ruler thead th",
        "(ths) => ths.map(th => (th.innerText || '').trim()).filter(Boolean)"
    )
    headers = [clean_text(h) for h in headers if clean_text(h)]
    if headers:
        return headers

    # Fallback: first non-empty tbody row as header
    header_row = page.eval_on_selector_all(
        "table.ruler tbody tr",
        r"""
        (trs) => {
          for (const tr of trs) {
            const tds = Array.from(tr.querySelectorAll("td"));
            const vals = tds.map(td => (td.innerText || "").trim()).filter(Boolean);
            if (vals.length >= 4) return vals; // heuristically "header-like"
          }
          return [];
        }
        """
    )
    header_row = [clean_text(h) for h in header_row if clean_text(h)]
    if header_row:
        return header_row

    # Last resort fallback (matches your screenshot)
    return ["Player", "Member ID", "Year of birth", "Play County", "Recent Form", "Tournaments"]


def extract_table_rows_u9_u10(page, logger, expected_cols: int):
    if page.locator("table.ruler").count() == 0:
        logger.info("No rankings table found on this page.")
        return []

    row_sel = "table.ruler tbody tr"
    if page.locator(row_sel).count() == 0:
        logger.info("No table rows found on this page.")
        return []

    raw_rows = page.eval_on_selector_all(
        row_sel,
        r"""(trs) => trs.map(tr => {
            const tds = Array.from(tr.querySelectorAll("td"));
            return tds.map(td => {
                const a = td.querySelector("a");
                const txt = (a ? a.textContent : td.textContent) || "";
                return txt.replace(/\u00a0/g, " ").replace(/\s+/g, " ").trim();
            });
        })"""
    )

    cleaned = []
    for r in raw_rows:
        r = [clean_text(x) for x in r]
        if any(r):
            cleaned.append(r)

    if not cleaned:
        return []

    # If the first row looks like headers (contains 'Member ID'), drop it
    first = [x.lower() for x in cleaned[0]]
    if any("member id" in x for x in first):
        cleaned = cleaned[1:]

    # Now filter to player rows by having a numeric member id somewhere
    data = []
    for r in cleaned:
        has_member_id = any(re.fullmatch(r"\d{6,12}", x) for x in r)
        if not has_member_id:
            continue

        if expected_cols > 0:
            if len(r) < expected_cols:
                r = r + [""] * (expected_cols - len(r))
            elif len(r) > expected_cols:
                r = r[:expected_cols]

        data.append(r)

    logger.info(f"Extracted {len(data)} rows before dedupe.")

    deduped = {}
    for r in data:
        member_ids = [x for x in r if re.fullmatch(r"\d{6,12}", x)]
        key = member_ids[0] if member_ids else tuple(r)
        deduped[key] = r

    data = list(deduped.values())

    logger.info(f"Returning {len(data)} rows after dedupe.")

    return data


def extract_u9_u10_rows(page, logger):
    rows = []

    # Wait for the table (same selector you inspected)
    table = page.locator("table.ruler")
    if not table.count():
        logger.info("No rankings table found on this page.")
        return rows

    for tr in table.locator("tbody tr").all():
        cells = [c.inner_text().strip() for c in tr.locator("td").all()]
        if not cells:
            continue

        # Expect exactly 6 columns for U9/U10
        if len(cells) != 6:
            logger.debug(f"Skipping malformed row: {cells}")
            continue

        rows.append(cells)

    logger.info(f"Extracted {len(rows)} rows (U9/U10).")
    return rows



def scrape_u9_u10_category(rank_id: int, category_id: int, sheet_tab: str, logger: logging.Logger):
    all_rows = []

    ranking_week = ""

    headers = [
        "Player",
        "Member ID",
        "Year of birth",
        "Play County",
        "Recent Form",
        "Tournaments",
        "Ranking Week",
    ]

    expected_cols = len(headers)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        for page_num in range(1, MAX_PAGES + 1):
            url = BASE_URL_TEMPLATE.format(
                rank_id=rank_id,
                category=category_id,
                p=page_num
            )

            logger.info(f"[{sheet_tab}] Fetching page {page_num}: {url}")

            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            accept_cookies_if_needed(page, logger)

# ✅ Wait for the table to exist (U9/U10 loads it after DOMContentLoaded)
            try:
                page.wait_for_selector("table.ruler", timeout=20000)

            
            except PlaywrightTimeoutError:
                logger.error(f"[{sheet_tab}] table.ruler did not appear on page {page_num}")
                save_debug_artifacts(page, page_num, logger, tag=f"{sheet_tab}_no_table")
                break

            if page_num == 1:
                ranking_week = get_ranking_week(page, logger)

            rows = extract_table_rows_u9_u10(page, logger, expected_cols-1)
            
            if not rows:
                logger.info(f"[{sheet_tab}] No rows on page {page_num}; stopping.")
                break

            rows = [r + [ranking_week] for r in rows]

            all_rows.extend(rows)

        context.close()
        browser.close()

    if not all_rows:
        raise RuntimeError(f"[{sheet_tab}] No rows scraped; refusing to overwrite sheet.")

#    write_rows_to_excel(
#        header=headers,
#        rows=all_rows,
#        sheet_tab=sheet_tab,
#        logger=logger,
#    )    

    write_rows_to_sheet_via_webapp(
        header=headers,
        rows=all_rows,
        sheet_tab=sheet_tab,
        clear_first=True,
        logger=logger,
    )



def main():
    logger = setup_logger()
    logger.info("Starting U9/U10 scrape")

    rank_id = resolve_recent_form_rank_id(logger)

    scrape_u9_u10_category(
        rank_id=rank_id,
        category_id=4660,      # U9
        sheet_tab="U9_rankings",
        logger=logger,
    )

    scrape_u9_u10_category(
        rank_id=rank_id,
        category_id=4658,      # U10
        sheet_tab="U10_rankings",
        logger=logger,
    )

    logger.info("U9_U10 Done")


if __name__ == "__main__":
    main()

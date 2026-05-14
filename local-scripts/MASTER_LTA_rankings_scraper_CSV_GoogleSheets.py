import csv
import logging
import os
import re
import requests
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


BASE_URL_TEMPLATE = (
    "https://competitions.lta.org.uk/ranking/category.aspx"
    "?id={rank_id}&category={category}&C{category}FOC=&C{category}RFPC=&p={p}&ps=100"
)


# Where to dump debugging artifacts (html + png) on failure
ARTIFACT_DIR = Path("lta_artifacts")

WEBAPP_URL = os.environ.get("LTA_SHEETS_WEBAPP_URL", "").strip()
WEBAPP_SECRET = os.environ.get("LTA_SHEETS_WEBAPP_SECRET", "").strip()

def write_rows_to_sheet_via_webapp(header, rows, sheet_tab="Rankings", clear_first=True, logger=None):
    if not WEBAPP_URL:
        raise RuntimeError("Missing env var LTA_SHEETS_WEBAPP_URL")

    payload = {
        "sheet": sheet_tab,
        "clearFirst": clear_first,
        "rows": [header] + rows,
    }

    params = {"secret": WEBAPP_SECRET} if WEBAPP_SECRET else None

    if logger:
        logger.info(f"Posting {len(rows)} rows to Google Sheet tab '{sheet_tab}'")

    r = requests.post(
        WEBAPP_URL,
        params=params,
        json=payload,
        timeout=120
    )
    r.raise_for_status()

    response_text = r.text.strip()
    if response_text.lower() != "ok":
        # Also accept JSON response with status "ok"
        try:
            import json
            parsed = json.loads(response_text)
            if parsed.get("status", "").lower() != "ok":
                raise RuntimeError(f"Web app returned: {r.text}")
            if logger:
                logger.info(f"Web app response: {response_text}")
        except (json.JSONDecodeError, AttributeError):
            raise RuntimeError(f"Web app returned: {r.text}")



def get_ranking_week(page, logger) -> str:
    """
    Returns the currently selected Ranking week, e.g. '6-2026'.
    Best source is the hidden <select name="...dlPublication...">.
    """
    sel = "select[name*='dlPublication']"
    try:
        if page.locator(sel).count():
            txt = page.eval_on_selector(
                sel,
                "s => (s.options[s.selectedIndex]?.textContent || '').trim()"
            )
            if txt:
                logger.info(f"Ranking week detected: {txt}")
                return txt
    except Exception:
        pass

    # Fallback: chosen widget (visible)
    try:
        fallback = page.locator("p:has(strong:has-text('Ranking week')) a.chosen-single span").first
        if fallback.count():
            txt = fallback.inner_text().strip()
            if txt:
                logger.info(f"Ranking week detected (fallback): {txt}")
                return txt
    except Exception:
        pass

    logger.warning("Could not detect Ranking week.")
    return ""




def write_rows_to_excel(header, rows, sheet_tab, logger, filename="LTA_MASTER_LIVE.xlsx"):
    excel_path = Path(__file__).resolve().parent / filename

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

    # Replace sheet if it already exists
    if sheet_tab in wb.sheetnames:
        ws = wb[sheet_tab]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_tab)

    ws.append(header)
    for r in rows:
        ws.append(r)

    wb.save(excel_path)
    logger.info(f"Wrote Excel tab '{sheet_tab}' in {excel_path}")




def setup_logger() -> logging.Logger:
    logger = logging.getLogger("lta_scrape")
    logger.setLevel(logging.INFO)

    # prevent duplicate handlers if you rerun in same interpreter
    if logger.handlers:
        return logger

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = f"lta_scrape_{ts}.log"

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)
    fh.setLevel(logging.INFO)

    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    sh.setLevel(logging.INFO)

    logger.addHandler(fh)
    logger.addHandler(sh)

    logger.info(f"Logging to: {log_file}")
    return logger


def clean_text(s: str) -> str:
    if s is None:
        return ""
    # normalize whitespace and strip NBSP etc
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def accept_cookies_if_needed(page, logger: logging.Logger) -> None:
    """
    Handles both:
    1) Cookie wall redirect: /cookiewall?ReturnUrl=...
    2) In-page cookie modal with an ACCEPT button (various implementations)
    """
    # 1) Cookie wall (URL-based)
    if "cookiewall" in page.url.lower():
        logger.info("Cookie wall detected via URL. Attempting to click ACCEPT...")
        # The cookie wall in your screenshot has a visible "ACCEPT" button
        # Try a few common selectors:
        candidates = [
            "button:has-text('ACCEPT')",
            "button:has-text('Accept')",
            "text=ACCEPT",
            "text=Accept",
        ]
        for sel in candidates:
            loc = page.locator(sel).first
            try:
                if loc.is_visible(timeout=1500):
                    loc.click()
                    page.wait_for_load_state("domcontentloaded")
                    logger.info(f"Clicked cookie wall ACCEPT using selector: {sel}")
                    break
            except Exception:
                pass

        logger.info(f"After cookie wall handling, current URL: {page.url}")

    # 2) In-page cookie modal (OneTrust / generic)
    # Don’t block if it isn't there; just best-effort click.
    modal_candidates = [
        "#onetrust-accept-btn-handler",
        "button#onetrust-accept-btn-handler",
        "button:has-text('ACCEPT')",
        "button:has-text('Accept')",
    ]
    for sel in modal_candidates:
        loc = page.locator(sel).first
        try:
            if loc.is_visible(timeout=1200):
                loc.click()
                logger.info(f"Clicked cookie modal ACCEPT using selector: {sel}")
                # Some modals animate away; no navigation needed
                break
        except Exception:
            pass


def extract_table_rows(page, logger):
    """
    Extract rankings from table.ruler using fixed td positions.
    Drops known non-data columns (movement + spacer columns) WITHOUT shifting.
    """
    row_selector = "table.ruler tbody tr:has(td.rank)"
    # If there are no rows on this page, return [] (end of pagination)
    if page.locator(row_selector).count() == 0:
        logger.info("No ranking rows found on this page.")
        return []


    rows = page.eval_on_selector_all(
        row_selector,
        r"""(trs) => trs.map(tr => {
            const tds = Array.from(tr.querySelectorAll("td"));

            // Expect 15 tds per row. If LTA changes layout, we want to know.
            // Return raw count too for debugging.
            const cellText = tds.map((td, idx) => {
                const a = td.querySelector("a");
                const txt = (a ? a.textContent : td.textContent) || "";
                return txt.replace(/\u00a0/g, " ").replace(/\s+/g, " ").trim();
            });

            return cellText;
        })"""
    )

    cleaned_rows = []
    for r in rows:
        # skip empty/blank rows if any appear
        if not any(r):
            continue

        # Drop columns by index: movement + spacers
        # Keep: 0,3,5..14
        out = [r[0], r[3], *r[5:15]]
        cleaned_rows.append(out)

    logger.info(f"Extracted {len(cleaned_rows)} rows (aligned).")
    return cleaned_rows



def get_headers(page):
    """
    Best effort: grab header cells from the table.
    If header parsing fails, fall back to expected columns.
    """
    # Try normal thead first
    headers = page.eval_on_selector_all(
        "table.ruler thead th",
        "(ths) => ths.map(th => (th.innerText || '').trim()).filter(Boolean)"
    )
    headers = [clean_text(h) for h in headers if clean_text(h)]

    # If the page uses a non-standard header layout, fall back.
    if not headers:
        headers = [
            "Rank",
            "Player",
            "Member ID",
            "Year of birth",
            "WTN Singles",
            "WTN Doubles",
            "Play County",
            "Singles Points",
            "Doubles points",
            "Tournaments",
            "Tournaments used for this calculation",
            "Total points",
        ]


    if not headers or headers[-1] != "Ranking Week":
        headers = headers + ["Ranking Week"]


    return headers


def resolve_current_rank_id(logger: logging.Logger) -> int:
    """
    Use the UI flow to discover the current weekly 'id' used in category.aspx URLs.
    Flow:
      1) https://competitions.lta.org.uk/ranking/
      2) click "LTA Combined Rankings"
      3) click "Open Male"
      4) parse id=NNNNN from resulting URL
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        logger.info("Resolving current weekly rank_id via UI flow...")

        page.goto("https://competitions.lta.org.uk/ranking/", wait_until="domcontentloaded", timeout=60000)
        accept_cookies_if_needed(page, logger)

        # Click "LTA Combined Rankings" (your DOM shows href contains ranking.aspx?rid=301)
        try:
            page.locator("a[href*='ranking.aspx?rid=301']").first.click()
        except Exception:
            # fallback if href changes but text stays
            page.locator("a:has-text('LTA Combined Rankings')").first.click()

        page.wait_for_load_state("domcontentloaded")
        accept_cookies_if_needed(page, logger)

        # Click "Open Male" (this leads to category.aspx?id=XXXXX&category=4544)
        page.locator("a:has-text('Open Male')").first.click()
        page.wait_for_load_state("domcontentloaded")

        final_url = page.url
        logger.info(f"Resolved URL: {final_url}")

        qs = parse_qs(urlparse(final_url).query)
        if "id" not in qs or not qs["id"] or not qs["id"][0].isdigit():
            raise RuntimeError(f"Could not parse rank_id from URL: {final_url}")

        rank_id = int(qs["id"][0])
        logger.info(f"Resolved rank_id={rank_id}")

        context.close()
        browser.close()

        return rank_id





def scrape_rankings_to_sheet(rank_id: int, category_id: int, sheet_tab: str, logger):

    all_rows = []
    ranking_week = ""


    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)  # keep identical to your current script
        context = browser.new_context()
        page = context.new_page()

        headers = None

        for page_num in range(1, 300):
            url = BASE_URL_TEMPLATE.format(rank_id=rank_id, category=category_id, p=page_num)

            logger.info(f"[{sheet_tab}] Fetching page {page_num}: {url}")

            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            accept_cookies_if_needed(page, logger)

            # Grab headers once, from the live table, using your existing function
            if headers is None:
                headers = get_headers(page)
                ranking_week = get_ranking_week(page, logger)


            rows = extract_table_rows(page, logger)

            if rows and ranking_week:
                rows = [r + [ranking_week] for r in rows]
            elif rows:
                rows = [r + [""] for r in rows]


            # Stop when we hit the first empty page (end of pagination)
            if not rows:
                logger.info(f"[{sheet_tab}] No rows on page {page_num}; assuming end of rankings. Stopping.")
                break

            all_rows.extend(rows)

        context.close()
        browser.close()

    logger.info(f"[{sheet_tab}] Total extracted rows: {len(all_rows)}")

    write_rows_to_sheet_via_webapp(
        header=headers,
        rows=all_rows,
        sheet_tab=sheet_tab,
        clear_first=True,
        logger=logger,
    )


    # write_rows_to_excel(headers, all_rows, sheet_tab, logger)




def save_debug_artifacts(page, page_num: int, logger: logging.Logger) -> None:
    ARTIFACT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = ARTIFACT_DIR / f"{ts}_page_{page_num}_fail.html"
    png_path = ARTIFACT_DIR / f"{ts}_page_{page_num}_fail.png"

    try:
        html_path.write_text(page.content(), encoding="utf-8")
        logger.info(f"Saved HTML: {html_path}")
    except Exception as e:
        logger.info(f"Failed to save HTML: {e}")

    try:
        page.screenshot(path=str(png_path), full_page=True)
        logger.info(f"Saved screenshot: {png_path}")
    except Exception as e:
        logger.info(f"Failed to save screenshot: {e}")




import subprocess
import sys
from pathlib import Path

def run_u9_u10_script(logger):
    script_path = Path(__file__).with_name(
        "LTA_rankings_scraper_U9_U10.py"
    )
    logger.info(f"Launching U9/U10 script: {script_path}")

    result = subprocess.run(
        [sys.executable, str(script_path)]
    )

    if result.returncode != 0:
        raise RuntimeError(
            f"U9/U10 script failed with exit code {result.returncode}"
        )


def main():
    logger = setup_logger()
    logger.info("Starting scrape...")

    rank_id = resolve_current_rank_id(logger)

    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4574, sheet_tab="U11_rankings", logger=logger)
    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4552, sheet_tab="U12_rankings", logger=logger)
    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4550, sheet_tab="U14_rankings", logger=logger)
    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4548, sheet_tab="U16_rankings", logger=logger)
    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4546, sheet_tab="U18_rankings", logger=logger)

    run_u9_u10_script(logger)

    scrape_rankings_to_sheet(rank_id=rank_id, category_id=4544, sheet_tab="Open_rankings", logger=logger)

    logger.info("Scraping Done.")

        
if __name__ == "__main__":
    main()
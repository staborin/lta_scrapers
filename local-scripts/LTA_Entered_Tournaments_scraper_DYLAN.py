"""
LTA Tournament Scraper
Scrapes tournament entries from competitions.lta.org.uk and exports to Excel.

Key features:
- Optional automated login using env vars:
    DYLAN_LTA_USERNAME
    DYLAN_LTA_PASSWORD
- Scrapes "My entries" from the home page after login
- For each tournament:
    - Scrapes Fact Sheet (status/grade/draw size)
    - Scrapes entries from Events tab (prefers "Entries" over "Online entries")
- Writes results by CLEARING the target worksheet tab each run, then writing fresh scraped data.
    - No UPSERT / no existing-tournament checks
    - Existing data on the worksheet is deleted before writing
"""

import os
import re
import logging
import requests
from datetime import datetime, timedelta

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

OUTPUT_XLSX = r"C:\Users\serge\OneDrive\A_Serge_Private\AI_Stuff\LTA\Ranking_Scrapers\LTA_MASTER_LIVE.xlsx"
OUTPUT_SHEET = "DYLAN_TOURNAMENTS"


WEBAPP_URL = os.environ.get("LTA_SHEETS_WEBAPP_URL", "").strip()
WEBAPP_SECRET = os.environ.get("LTA_SHEETS_WEBAPP_SECRET", "").strip()

def write_wide_to_sheet_via_webapp(tournaments, sheet_tab, clear_first, logger: logging.Logger):
    """
    Write tournaments to Google Sheets in the same wide/column-block format as Excel.

    Layout (mirrors write_fresh_to_excel exactly):
      Row 1: "Tournament: <name>"       (col A of each block)
      Row 2: date value                 (col B of each block)
      Row 3: event value                (col B of each block)
      Row 4: status value               (col B of each block)
      Row 5: grade value                (col B of each block)
      Row 6: draw_size value            (col B of each block)
      Row 7: blank
      Row 8: "Entry Name" | "Entry Date" (headers)
      Row 9+: entries

    Each tournament occupies 2 columns + 1 blank spacer column.
    """
    if not WEBAPP_URL:
        raise RuntimeError("Missing env var LTA_SHEETS_WEBAPP_URL")

    if not tournaments:
        logger.info("No tournaments to write — skipping Google Sheets update.")
        return

    num_tournaments = len(tournaments)
    total_cols = num_tournaments * 3  # 2 data cols + 1 spacer per tournament

    max_entries = max((len(t.get("entries") or []) for t in tournaments), default=0)
    total_rows = 9 + max_entries  # row 7 now holds URL

    grid = [[None] * total_cols for _ in range(total_rows)]

    for t_idx, tournament in enumerate(tournaments):
        col = t_idx * 3  # 0-indexed column for this tournament's left column

        grid[0][col] = f"Tournament: {tournament.get('name', '')}"
        grid[0][col + 1] = tournament.get("venue", "")

        grid[1][col] = "Date:"
        grid[1][col + 1] = tournament.get("date", "")
        grid[1][col + 2] = tournament.get("start_time", "")

        grid[2][col] = "Event:"
        grid[2][col + 1] = tournament.get("event", "")

        grid[3][col] = "Status:"
        grid[3][col + 1] = tournament.get("status", "")
        grid[3][col + 2] = tournament.get("closing_date", "")

        grid[4][col] = "Grade:"
        grid[4][col + 1] = tournament.get("grade", "")

        grid[5][col] = "Draw Size:"
        grid[5][col + 1] = tournament.get("draw_size", "")

        grid[6][col] = "URL:"
        grid[6][col + 1] = tournament.get("link", "")  # entries URL

        grid[8][col] = "Entry Name"
        grid[8][col + 1] = "Entry Date"

        for e_idx, entry in enumerate(tournament.get("entries") or []):
            grid[9 + e_idx][col] = entry.get("name", "")
            grid[9 + e_idx][col + 1] = entry.get("entry_date", "")

    # Replace None with "" for JSON serialisation
    grid = [["" if cell is None else cell for cell in row] for row in grid]

    url = WEBAPP_URL
    if WEBAPP_SECRET:
        joiner = "&" if "?" in url else "?"
        url = f"{url}{joiner}secret={WEBAPP_SECRET}"

    payload = {
        "sheet": sheet_tab,
        "clearFirst": bool(clear_first),
        "rows": grid,
    }

    logger.info(f"Posting wide-format grid ({total_rows} rows x {total_cols} cols) "
                f"to Google Sheet tab '{sheet_tab}'")
    r = requests.post(url, json=payload, timeout=120)
    r.raise_for_status()
    logger.info(f"Google Sheets response: {r.text.strip()}")


def setup_logger() -> logging.Logger:
    logger = logging.getLogger("lta_tournaments")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        ch.setFormatter(fmt)
        logger.addHandler(ch)
    return logger


def parse_tournament_end_date(date_str: str) -> date | None:
    """
    Accepts:
      - "15/02/2026 to 21/02/2026"
      - "15/02/2026 - 21/02/2026"
      - "15/02/2026"
    Returns the END date if a range, else the single date.
    """
    if not date_str:
        return None
    dates = re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", date_str)
    if not dates:
        return None
    try:
        return datetime.strptime(dates[-1], "%d/%m/%Y").date()  # <- end date if range
    except ValueError:
        return None

def parse_closing_deadline(closing_str: str) -> datetime | None:
    """
    Accepts:
      - "05/02/2026 10:00"
    Returns a datetime or None.
    """
    if not closing_str:
        return None
    closing_str = closing_str.strip()
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H.%M"):
        try:
            return datetime.strptime(closing_str, fmt)
        except ValueError:
            pass
    return None







# ---------------------------
# Date parsing & status logic
# ---------------------------

def parse_date(date_string: str):
    """Parse various date formats from LTA site. If date range, use first date."""
    if not date_string:
        return None

    s = date_string.strip()

    # Normalize some range separators (en dash/em dash)
    s = s.replace("–", "-").replace("—", "-")

    try:
        # Single date format: 22/02/2026
        if "/" in s and "to" not in s.lower() and "-" not in s:
            return datetime.strptime(s, "%d/%m/%Y").date()

        # Range formats: "15/02/2026-21/02/2026" or "15/02/2026 to 21/02/2026"
        if "to" in s.lower() or "-" in s:
            if "to" in s.lower():
                first_date = re.split(r"\bto\b", s, flags=re.IGNORECASE)[0].strip()
            else:
                first_date = s.split("-")[0].strip()
            return datetime.strptime(first_date, "%d/%m/%Y").date()

    except Exception as e:
        print(f"Error parsing date '{date_string}': {e}")

    return None


def extract_fact_sheet_value(page, event_text, label_text):
    """
    Extracts a dt/dd value from the Fact Sheet for a specific event.

    Universal approach:
    1) Try accordion structure: find event toggle -> expand -> read label inside that panel
    2) Fallback: read label inside the first visible dl.list--flex (non-accordion pages)

    Returns: string or None
    """
    event_text = (event_text or "").strip()
    label_text = (label_text or "").strip()

    # --- A) Accordion path (Croydon-style) ---
    try:
        toggle = page.locator("dt.list__label button").filter(
            has=page.locator("span.text--bold", has_text=re.compile(rf"^\s*{re.escape(event_text)}\s*$", re.I))
        ).first

        if toggle.count() > 0:
            expanded = toggle.get_attribute("aria-expanded")
            if expanded != "true":
                toggle.click()
                page.wait_for_timeout(500)

            dt_node = toggle.locator("xpath=ancestor::dt[1]").first
            panel = dt_node.locator("xpath=following-sibling::dd[1]").first

            dt_label = panel.locator("dt.list__label", has_text=re.compile(rf"^\s*{re.escape(label_text)}\s*$", re.I)).first
            if dt_label.count() > 0:
                dd_val = dt_label.locator("xpath=following-sibling::dd[1]").first
                val = dd_val.inner_text().strip()
                return val or None
    except Exception:
        pass

    # --- B) Fallback path (non-accordion pages) ---
    try:
        container = page.locator("dl.list--flex").first
        if container.count() == 0:
            return None

        dt_label = container.locator("dt.list__label", has_text=re.compile(rf"^\s*{re.escape(label_text)}\s*$", re.I)).first
        if dt_label.count() == 0:
            return None

        dd_val = dt_label.locator("xpath=following-sibling::dd[1]").first
        val = dd_val.inner_text().strip()
        return val or None
    except Exception:
        return None



def parse_closing_deadline(closing_str: str) -> datetime | None:
    """
    Robustly parse closing deadline from strings like:
      - "05/02/2026 10:00"
      - "05/02/2026 10:00 GMT"
      - "Closing deadline 05/02/2026 10:00"
      - "05/02/2026 10:00:00"
      - strings containing NBSP / extra whitespace
    Returns datetime or None.
    """
    if not closing_str:
        return None

    s = closing_str.replace("\u00a0", " ").strip()  # NBSP -> space

    # Pull out the first "dd/mm/yyyy HH:MM" (optionally with seconds)
    m = re.search(r"\b(\d{2}/\d{2}/\d{4})\s+(\d{1,2}:\d{2})(?::\d{2})?\b", s)
    if not m:
        return None

    dt_str = f"{m.group(1)} {m.group(2)}"  # ignore timezone words like GMT/BST
    try:
        return datetime.strptime(dt_str, "%d/%m/%Y %H:%M")
    except ValueError:
        return None


def parse_timings_to_event_date(timings_str: str | None, start_date_str: str | None) -> str | None:
    """
    Attempt to extract concrete event date(s) from the freeform "Timings" field,
    using the tournament "Start date" as an anchor for weekday-only references.

    Strategy order:
      1) Full dd/mm/yyyy dates found -> outer min-max range
      2) "day month" patterns like "28 March", "3rd April" -> infer year from start_date
      3) Weekday names only (Monday/Mon/Weds/etc.) -> next occurrence on or after start_date

    Returns: "dd/mm/yyyy" or "dd/mm/yyyy - dd/mm/yyyy", or None if unparseable.
    """
    if not timings_str or not timings_str.strip():
        return None

    timings = timings_str.strip()

    # Parse the start date as our anchor
    anchor_date = None
    if start_date_str:
        m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", start_date_str)
        if m:
            try:
                anchor_date = datetime.strptime(m.group(1), "%d/%m/%Y").date()
            except ValueError:
                pass

    # --- Strategy 1: Full dd/mm/yyyy dates in the Timings string ---
    full_dates = re.findall(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", timings)
    if full_dates:
        parsed = []
        for d in full_dates:
            try:
                parsed.append(datetime.strptime(d, "%d/%m/%Y").date())
            except ValueError:
                pass
        if parsed:
            mn, mx = min(parsed), max(parsed)
            if mn == mx:
                return mn.strftime("%d/%m/%Y")
            return f"{mn.strftime('%d/%m/%Y')} - {mx.strftime('%d/%m/%Y')}"

    # --- Strategy 2: "day month" patterns (e.g. "28 March", "3rd April", "29th June") ---
    MONTH_MAP = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12,
        "jan": 1, "feb": 2, "mar": 3, "apr": 4,
        "jun": 6, "jul": 7, "aug": 8,
        "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }

    day_month_pattern = r"\b(\d{1,2})(?:st|nd|rd|th)?\s+(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec)\b"
    day_month_matches = re.findall(day_month_pattern, timings, re.IGNORECASE)

    if day_month_matches:
        all_num_positions = list(re.finditer(r"\b(\d{1,2})(?:st|nd|rd|th)?\b", timings))
        month_positions = list(re.finditer(
            r"\b(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec)\b",
            timings, re.IGNORECASE
        ))

        already_captured = {(d, m.lower()) for d, m in day_month_matches}

        for num_m in all_num_positions:
            day_val = int(num_m.group(1))
            if day_val < 1 or day_val > 31:
                continue
            after_num = timings[num_m.end():num_m.end() + 15].strip().lower()
            directly_has_month = False
            for mname in MONTH_MAP:
                if after_num.startswith(mname):
                    directly_has_month = True
                    break
            if directly_has_month:
                continue

            for mp in month_positions:
                if mp.start() > num_m.end():
                    orphan_month = mp.group(1).lower()
                    if (str(day_val), orphan_month) not in already_captured:
                        day_month_matches.append((str(day_val), orphan_month))
                        already_captured.add((str(day_val), orphan_month))
                    break

    if day_month_matches and anchor_date:
        parsed = []
        year = anchor_date.year
        for day_str, month_str in day_month_matches:
            month_num = MONTH_MAP.get(month_str.lower())
            if not month_num:
                continue
            try:
                d = datetime(year, month_num, int(day_str)).date()
                if d.month < anchor_date.month - 2:
                    d = datetime(year + 1, month_num, int(day_str)).date()
                parsed.append(d)
            except ValueError:
                continue
        if parsed:
            mn, mx = min(parsed), max(parsed)
            if mn == mx:
                return mn.strftime("%d/%m/%Y")
            return f"{mn.strftime('%d/%m/%Y')} - {mx.strftime('%d/%m/%Y')}"

    # --- Strategy 3: Weekday names only ---
    WEEKDAY_MAP = {
        "monday": 0, "mon": 0,
        "tuesday": 1, "tue": 1, "tues": 1,
        "wednesday": 2, "wed": 2, "weds": 2,
        "thursday": 3, "thu": 3, "thur": 3, "thurs": 3,
        "friday": 4, "fri": 4,
        "saturday": 5, "sat": 5,
        "sunday": 6, "sun": 6,
    }

    weekday_names = sorted(WEEKDAY_MAP.keys(), key=len, reverse=True)
    weekday_pattern = r"\b(" + "|".join(weekday_names) + r")\b"
    weekday_matches = re.findall(weekday_pattern, timings, re.IGNORECASE)

    if weekday_matches and anchor_date:
        target_weekdays = set()
        for w in weekday_matches:
            wnum = WEEKDAY_MAP.get(w.lower())
            if wnum is not None:
                target_weekdays.add(wnum)

        range_pattern = r"\b(" + "|".join(weekday_names) + r")\s*[-–—]\s*(" + "|".join(weekday_names) + r")\b"
        range_matches = re.findall(range_pattern, timings, re.IGNORECASE)
        for start_day, end_day in range_matches:
            s = WEEKDAY_MAP.get(start_day.lower())
            e = WEEKDAY_MAP.get(end_day.lower())
            if s is not None and e is not None:
                d = s
                while True:
                    target_weekdays.add(d)
                    if d == e:
                        break
                    d = (d + 1) % 7

        if target_weekdays:
            dates = []
            for wd in target_weekdays:
                days_ahead = (wd - anchor_date.weekday()) % 7
                dates.append(anchor_date + timedelta(days=days_ahead))
            dates.sort()
            mn, mx = dates[0], dates[-1]
            if mn == mx:
                return mn.strftime("%d/%m/%Y")
            return f"{mn.strftime('%d/%m/%Y')} - {mx.strftime('%d/%m/%Y')}"

    return None


def extract_start_time(timings_str: str | None) -> str | None:
    """
    Extract a start time from a Timings string if present.
    Matches patterns like: "1.45pm start", "10am", "2:30pm", "10.00am start"
    Returns the time string (e.g. "1.45pm") or None.
    """
    if not timings_str:
        return None
    m = re.search(r'\b(\d{1,2}(?:[.:]\d{2})?\s*[ap]m)\b', timings_str, re.IGNORECASE)
    if m:
        return m.group(1).strip().lower()
    return None


def determine_status(closing_deadline_str: str | None, tournament_date_str: str | None) -> str:
    """
    Rules:
      - If tournament end date is in the past -> Played
      - Else if closing deadline is in the future -> Open
      - Else if closing deadline is in the past -> Closed
      - Else -> Unknown
    """
    today = datetime.now().date()

    end_date = parse_tournament_end_date(tournament_date_str or "")
    if end_date and end_date < today:
        return "Played"

    closing_dt = parse_closing_deadline(closing_deadline_str or "")
    if closing_dt:
        if datetime.now() < closing_dt:
            return "Open"
        return "Closed"

    # If we can’t parse either, we can’t be confident
    return "Unknown"


# ---------------------------
# Helpers
# ---------------------------

def clean_tournament_name(name: str) -> str:
    """Remove 'London & South East Tour - ' prefix (as per your original logic)."""
    return name.replace("London & South East Tour - ", "").strip()


def dismiss_cookie_banner(page):
    """Dismiss cookie banner if present."""
    try:
        cookie_accept = page.query_selector('button:has-text("ACCEPT"), button:has-text("Accept")')
        if cookie_accept:
            cookie_accept.click()
            page.wait_for_timeout(800)
            print("✓ Cookie banner dismissed")
    except Exception as e:
        print(f"(Cookie banner dismiss warning: {e})")


def login_if_needed(page):
    username = os.environ.get("DYLAN_LTA_USERNAME", "").strip()
    password = os.environ.get("DYLAN_LTA_PASSWORD", "").strip()

    if not username or not password:
        print("\n⏸️  Env vars not set: DYLAN_LTA_USERNAME / DYLAN_LTA_PASSWORD")
        print("Please log in manually.")
        input("Press Enter once you're logged in and on the home page...")
        return

    print("\nAttempting automated login...")

    # 1) Click top-right "Log in" (on competitions site)
    try:
        loc = page.get_by_role("link", name=re.compile(r"^\s*log\s*in\s*$", re.I))
        if loc.count() == 0:
            loc = page.get_by_role("button", name=re.compile(r"^\s*log\s*in\s*$", re.I))
        if loc.count() > 0:
            loc.first.click(timeout=8000)
            page.wait_for_load_state("domcontentloaded", timeout=15000)
            print("✓ Clicked top-right 'Log in'")
    except PlaywrightTimeoutError:
        print("(Top-right 'Log in' not found/clickable — may already be logged in.)")
    except Exception as e:
        print(f"(Top-right 'Log in' click warning: {e})")

    # 2) Click member "LOG IN" and CAPTURE where it navigates (same tab or popup)
    login_page = page  # default: same tab

    try:
        page.wait_for_load_state("domcontentloaded", timeout=15000)

        member_login_btn = page.get_by_role("button", name=re.compile(r"^\s*log\s*in\s*$", re.I))
        if member_login_btn.count() == 0:
            member_login_btn = page.locator('a.btn.btn-primary', has_text=re.compile(r"^\s*log\s*in\s*$", re.I))

        # Try popup first (OAuth flows sometimes open new tab)
        popup = None
        try:
            with page.expect_popup(timeout=4000) as pop:
                member_login_btn.first.click(timeout=15000)
            popup = pop.value
            login_page = popup
            print("✓ Member 'LOG IN' opened a popup/new tab")
        except Exception:
            # No popup -> same tab navigation
            member_login_btn.first.click(timeout=15000)
            login_page = page
            print("✓ Clicked member 'LOG IN' (same tab)")

        login_page.wait_for_load_state("domcontentloaded", timeout=20000)
        print(f"  Login page URL now: {login_page.url}")

    except Exception as e:
        print(f"(Member 'LOG IN' click warning: {e})")
        login_page = page  # fall back

    # 3) Fill Salesforce LWC login form (on the CORRECT page)
    try:
        print("  Waiting for LTA username/password inputs...")

        # Wait until we’re actually on the Salesforce login host/page
        # (If this never matches, it still continues; it’s just a guard)
        try:
            login_page.wait_for_timeout(500)
        except Exception:
            pass

        username_input = login_page.locator('input[placeholder="Username"][type="text"]')
        password_input = login_page.locator('input[placeholder="Password"][type="password"]')

        # Before waiting, print counts to avoid silent hangs
        print(f"  username_input count: {username_input.count()}")
        print(f"  password_input count: {password_input.count()}")

        # Use wait_for_selector as a stronger sync point
        login_page.wait_for_selector('input[placeholder="Username"][type="text"]', state="visible", timeout=30000)
        login_page.wait_for_selector('input[placeholder="Password"][type="password"]', state="visible", timeout=30000)

        # Re-resolve after waits (LWC sometimes re-renders)
        username_input = login_page.locator('input[placeholder="Username"][type="text"]').first
        password_input = login_page.locator('input[placeholder="Password"][type="password"]').first

        # Make sure they are interactable
        username_input.scroll_into_view_if_needed()
        password_input.scroll_into_view_if_needed()

        # Lightning: click + fill is often OK, but if it’s flaky, use keyboard typing
        username_input.click(force=True)
        username_input.fill("")  # clear
        login_page.keyboard.type(username, delay=30)

        password_input.click(force=True)
        password_input.fill("")  # clear
        login_page.keyboard.type(password, delay=30)

        print("  ✓ Credentials typed")

        login_button = login_page.locator('button[title="Log in"]').first
        login_button.wait_for(state="visible", timeout=15000)
        login_button.click()

        # DO NOT trust a single load_state here — the redirect chain can still be in flight.
        print("  Waiting for post-login redirect + authenticated competitions session...")
        wait_for_competitions_logged_in(page, timeout_ms=90000)
        print("✓ Logged in successfully (competitions session active)")


        # --- IMPORTANT: return to competitions site after Salesforce login ---
        try:
            # If login happened in a popup, close it and use the original page
            if login_page is not page:
                try:
                    login_page.close()
                except Exception:
                    pass

            # Always re-anchor on competitions site (prevents staying on mylta.my.site.com)
            page.goto("https://competitions.lta.org.uk/", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1200)

            # Confirm the left nav is present (your DOM shows title="Tournaments")
            page.wait_for_selector('a.primary-nav__link[title="Tournaments"]', timeout=20000)

            print("✓ Back on competitions site and nav is present")
        except Exception as e:
            print(f"⚠️  After-login re-anchor failed: {e}")



    except Exception as e:
        print("⚠️  Credential entry failed")
        print(e)
        try:
            login_page.screenshot(path="login_lwc_error.png", full_page=True)
            with open("login_lwc_error.html", "w", encoding="utf-8") as f:
                f.write(login_page.content())
            print("Saved login_lwc_error.png / login_lwc_error.html")
            print(f"URL at failure: {login_page.url}")
        except Exception:
            pass


# ---------------------------
# Excel: CLEAR + WRITE FRESH
# ---------------------------

def clear_worksheet(ws):
    """Clear the worksheet completely (values + styles) by deleting all rows/cols."""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column and ws.max_column > 0:
        ws.delete_cols(1, ws.max_column)


def write_fresh_to_excel(tournaments, filename=OUTPUT_XLSX, sheet_name=OUTPUT_SHEET):
    """
    Clears the target worksheet tab and writes the newly scraped tournaments
    in horizontal blocks (2 cols + 1 spacer), matching the existing layout.
    """
    if not os.path.exists(filename):
        raise RuntimeError(f"Excel file does not exist: {filename}")

    wb = load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' does not exist in workbook")

    ws = wb[sheet_name]

    # Clear everything on the sheet
    clear_worksheet(ws)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    _center_align = Alignment(horizontal="center", vertical="center")  # retained, if you want later

    current_col = 1

    for tournament in tournaments:
        # Header row
        ws.cell(row=1, column=current_col, value=f"Tournament: {tournament.get('name','')}")
        ws.cell(row=1, column=current_col).font = Font(bold=True, size=12)

        # Metadata
        ws.cell(row=2, column=current_col, value="Date:")
        ws.cell(row=2, column=current_col + 1, value=tournament.get("date", "N/A"))

        ws.cell(row=3, column=current_col, value="Event:")
        ws.cell(row=3, column=current_col + 1, value=tournament.get("event", "N/A"))

        ws.cell(row=4, column=current_col, value="Status:")
        ws.cell(row=4, column=current_col + 1, value=tournament.get("status", "N/A"))

        ws.cell(row=5, column=current_col, value="Grade:")
        ws.cell(row=5, column=current_col + 1, value=tournament.get("grade", "N/A"))

        ws.cell(row=6, column=current_col, value="Draw Size:")
        ws.cell(row=6, column=current_col + 1, value=tournament.get("draw_size", "N/A"))

        # Entries header
        ws.cell(row=8, column=current_col, value="Entry Name").font = header_font
        ws.cell(row=8, column=current_col).fill = header_fill
        ws.cell(row=8, column=current_col + 1, value="Entry Date").font = header_font
        ws.cell(row=8, column=current_col + 1).fill = header_fill

        # Entries
        row = 9
        for entry in (tournament.get("entries") or []):
            ws.cell(row=row, column=current_col, value=entry.get("name", ""))
            ws.cell(row=row, column=current_col + 1, value=entry.get("entry_date", ""))
            row += 1

        # Column widths
        ws.column_dimensions[ws.cell(row=1, column=current_col).column_letter].width = 25
        ws.column_dimensions[ws.cell(row=1, column=current_col + 1).column_letter].width = 25

        # Next block (2 cols + 1 spacer)
        current_col += 3

    wb.save(filename)
    wb.close()
    print(f"Excel file refreshed (cleared + rewritten): {filename}")

# ---------------------------
# Excel: legacy UPSERT helpers (unused)
# ---------------------------

def load_existing_tournaments(filename=OUTPUT_XLSX, sheet_name=OUTPUT_SHEET):
    """
    Return mapping: {tournament_name: start_col} for existing blocks
    from the specified worksheet.
    """
    if not os.path.exists(filename):
        return {}

    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' does not exist")

        ws = wb[sheet_name]
        existing = {}

        col = 1
        while col <= ws.max_column:
            cell_value = ws.cell(row=1, column=col).value
            if isinstance(cell_value, str) and cell_value.startswith("Tournament:"):
                tourn_name = cell_value.replace("Tournament:", "").strip()
                existing[tourn_name] = col
            col += 3

        wb.close()
        return existing

    except Exception as e:
        print(f"Error loading existing tournaments: {e}")
        return {}



def upsert_to_excel(tournaments, filename=OUTPUT_XLSX, sheet_name=OUTPUT_SHEET):
    """
    Update existing workbook + specific worksheet.
    Updates existing tournament blocks or appends new ones.
    """

    if not os.path.exists(filename):
        raise RuntimeError(f"Excel file does not exist: {filename}")

    wb = load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' does not exist in workbook")

    ws = wb[sheet_name]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    _center_align = Alignment(horizontal="center", vertical="center")  # retained, if you want later

    existing_map = load_existing_tournaments(filename, sheet_name) if os.path.exists(filename) else {}

    def next_free_col():
        # Align to 3-col blocks (2 data cols + 1 separator)
        if ws.max_column < 1:
            return 1
        last_block_start = ((ws.max_column - 1) // 3) * 3 + 1
        return last_block_start + 3

    def clear_old_entries(start_col):
        # Clear entries area (row 9 down) for the 2 data columns
        max_row = max(ws.max_row, 9)
        for r in range(9, max_row + 1):
            ws.cell(row=r, column=start_col).value = None
            ws.cell(row=r, column=start_col + 1).value = None

    for tournament in tournaments:
        start_col = existing_map.get(tournament["name"])
        if not start_col:
            start_col = next_free_col()
            existing_map[tournament["name"]] = start_col

        # Header row
        ws.cell(row=1, column=start_col, value=f"Tournament: {tournament['name']}")
        ws.cell(row=1, column=start_col).font = Font(bold=True, size=12)

        # Metadata
        ws.cell(row=2, column=start_col, value="Date:")
        ws.cell(row=2, column=start_col + 1, value=tournament.get("date", "N/A"))

        ws.cell(row=3, column=start_col, value="Event:")
        ws.cell(row=3, column=start_col + 1, value=tournament.get("event", "N/A"))

        ws.cell(row=4, column=start_col, value="Status:")
        ws.cell(row=4, column=start_col + 1, value=tournament.get("status", "N/A"))

        ws.cell(row=5, column=start_col, value="Grade:")
        ws.cell(row=5, column=start_col + 1, value=tournament.get("grade", "N/A"))

        ws.cell(row=6, column=start_col, value="Draw Size:")
        ws.cell(row=6, column=start_col + 1, value=tournament.get("draw_size", "N/A"))

        # Entries header
        ws.cell(row=8, column=start_col, value="Entry Name").font = header_font
        ws.cell(row=8, column=start_col).fill = header_fill
        ws.cell(row=8, column=start_col + 1, value="Entry Date").font = header_font
        ws.cell(row=8, column=start_col + 1).fill = header_fill

        # Replace entries
        clear_old_entries(start_col)
        for i, entry in enumerate(tournament.get("entries", []), start=9):
            ws.cell(row=i, column=start_col, value=entry.get("name", ""))
            ws.cell(row=i, column=start_col + 1, value=entry.get("entry_date", ""))

        # Column widths
        ws.column_dimensions[ws.cell(row=1, column=start_col).column_letter].width = 25
        ws.column_dimensions[ws.cell(row=1, column=start_col + 1).column_letter].width = 25

    wb.save(filename)
    wb.close()
    print(f"Excel file updated: {filename}")



def wait_for_competitions_logged_in(page, timeout_ms=60000):
    """
    Wait until we're back on competitions.lta.org.uk *and* the UI indicates an authenticated session.

    We avoid URL-only checks because you can be on competitions but still not have the logged-in session active.
    """
    import time
    start = time.time()

    # 1) Wait until we're on competitions domain (post-IdP redirect chain)
    page.wait_for_url(re.compile(r"^https://competitions\.lta\.org\.uk/.*", re.I), timeout=timeout_ms)

    # 2) Now wait until something that only appears when logged in is present.
    # Based on your DOM/screens:
    # - Left nav includes: a.primary-nav__link[title="Tournaments"]
    # - Logged-in pages often show "Log off" in footer/account area
    # We'll accept ANY of these signals.

    signals = [
        'a.primary-nav__link[title="Tournaments"]',
        'a.primary-nav__link[href="/tournaments"]',
        'text=/\\bLog off\\b/i',
        'a[href*="logoff"]',
        'text=/\\bMy tournaments\\b/i',  # when already on tournaments page and logged in
    ]

    last_err = None
    while (time.time() - start) * 1000 < timeout_ms:
        try:
            page.wait_for_load_state("domcontentloaded", timeout=10000)
            # networkidle is sometimes too strict, but helps when redirects are still happening
            try:
                page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass

            for sel in signals:
                try:
                    if page.locator(sel).count() > 0:
                        return
                except Exception:
                    continue
        except Exception as e:
            last_err = e

        page.wait_for_timeout(500)

    raise TimeoutError(f"Timed out waiting for logged-in state on competitions site. Last error: {last_err}")






# ---------------------------
# Scraping: "My entries"
# ---------------------------


def scrape_my_entries(page):
    """
    Scrape tournaments from the /tournaments page under "My tournaments".

    Returns: list of dicts with name/date/link. The entered event is discovered later
    by opening each tournament page and clicking "My Entry".
    """
    print("Scraping 'My tournaments' from /tournaments ...")
    tournaments = []

    def _wait_until_logged_in():
        """
        We need a reliable 'logged-in' signal before clicking Tournaments.
        On LTA/TournamentSoftware pages, a good signal is seeing 'Log off' in the footer/menu,
        or the user initials/avatar area. We'll wait for either.
        """
        print("Waiting for post-login state to settle...")

        # Try several "logged-in" signals; whichever appears first wins.
        # (These are intentionally broad and low-risk.)
        candidates = [
            page.locator("text=/\\bLog\\s*off\\b/i"),
            page.locator("a:has-text('Log off')"),
            page.locator("text=/\\bDylan\\b/i"),   # optional, harmless if not found
            page.locator("text=/\\bLuka\\b/i"),    # optional, harmless if not found
            page.locator("div.user, .user, .avatar, .profile"),  # generic
        ]

        # Give the app time to finish redirects/XHR after login
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(800)

        # Poll quickly up to ~20s for any signal
        for _ in range(40):
            for loc in candidates:
                try:
                    if loc.count() > 0:
                        # If it's there, we consider login "settled"
                        return True
                except Exception:
                    pass
            page.wait_for_timeout(500)

        # If we can't prove it, we still continue (better than hard-failing),
        # but we log it loudly.
        print("⚠️  Could not positively confirm logged-in state (continuing anyway).")
        return False

    try:
        print(f"Current URL before nav click: {page.url}")

        # 0) CRITICAL: wait for post-login to fully settle
        _wait_until_logged_in()

        # 1) Click the left-nav "Tournaments" item
        # Your DOM shows: <a href="/tournaments" title="Tournaments" class="primary-nav__link nav-link ...">
        print("Clicking left-nav: Tournaments...")

        nav_candidates = [
            page.locator('a.primary-nav__link.nav-link[href="/tournaments"]').first,
            page.locator('a.primary-nav__link[href="/tournaments"]').first,
            page.locator('a[href="/tournaments"][title="Tournaments"]').first,
            page.locator('a[href="/tournaments"]').first,
        ]

        clicked = False
        for nav in nav_candidates:
            try:
                if nav.count() == 0:
                    continue
                nav.scroll_into_view_if_needed()
                nav.click(timeout=15000, force=True)
                clicked = True
                break
            except Exception:
                continue

        if not clicked:
            raise RuntimeError("Could not find/click the 'Tournaments' nav link (<a href='/tournaments'>).")

        # 2) Wait until we're actually on /tournaments
        page.wait_for_url(re.compile(r".*/tournaments/?($|\?)", re.I), timeout=30000)
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(800)
        print(f"✓ Landed on: {page.url}")

        # 3) Find the "My tournaments" module (robust)
        # DOM shows: <div class="module module--card"> ... <span class="module__title--main">My tournaments</span>
        my_section = page.locator("div.module.module--card").filter(
            has=page.locator("span.module__title--main", has_text=re.compile(r"My tournaments", re.I))
        ).first

        # fallback: any module containing "My tournaments" text
        if my_section.count() == 0:
            my_section = page.locator("div.module").filter(
                has_text=re.compile(r"My tournaments", re.I)
            ).first

        if my_section.count() == 0:
            print("ERROR: Could not find 'My tournaments' section on /tournaments.")
            try:
                page.screenshot(path="debug_no_my_tournaments.png", full_page=True)
                with open("debug_no_my_tournaments.html", "w", encoding="utf-8") as f:
                    f.write(page.content())
                print("Saved debug_no_my_tournaments.png / debug_no_my_tournaments.html")
            except Exception:
                pass
            return tournaments

        print("✓ Found 'My tournaments' section")

        # 4) IMPORTANT: wait for the module to actually populate with tournament links (XHR)
        # Links are: a[href^="/sport/tournament"]
        links = my_section.locator('a[href^="/sport/tournament"]')
        print("Waiting for tournament cards to populate...")

        populated = False
        for _ in range(40):  # up to ~20s (40 * 500ms)
            try:
                if links.count() > 0:
                    populated = True
                    break
            except Exception:
                pass
            page.wait_for_timeout(500)

        if not populated:
            # No tournament links inside "My tournaments" — the player has no entered tournaments.
            # Do NOT fall back to a global page scan: that would pick up "Last visited" cards
            # which are unrelated tournaments the player browsed but did not enter.
            print("ℹ️  No tournament links found inside 'My tournaments'. Player has no entered tournaments.")
            return tournaments  # empty list — correct behaviour

        n = links.count()
        print(f"Found {n} tournament link(s)")

        if n == 0:
            print("ERROR: Still found 0 tournament links (even though UI may show them).")
            try:
                page.screenshot(path="debug_zero_tournament_links.png", full_page=True)
                with open("debug_zero_tournament_links.html", "w", encoding="utf-8") as f:
                    f.write(page.content())
                print("Saved debug_zero_tournament_links.png / debug_zero_tournament_links.html")
            except Exception:
                pass
            return tournaments

        seen = set()

        for i in range(n):
            a = links.nth(i)
            href = a.get_attribute("href") or ""
            if not href:
                continue

            if not href.startswith("http"):
                href = "https://competitions.lta.org.uk" + href

            if href in seen:
                continue
            seen.add(href)

            # Card container is often div.list__item (not necessarily li)
            card = a.locator("xpath=ancestor::*[contains(@class,'list__item')][1]").first
            card_text = (card.inner_text() or "") if card.count() > 0 else (a.inner_text() or "")

            # Name typically in h4.media__title
            name = ""
            if card.count() > 0 and card.locator("h4.media__title").count() > 0:
                name = clean_tournament_name((card.locator("h4.media__title").first.inner_text() or "").strip())
            else:
                # fallback: first non-empty line of card text
                lines = [ln.strip() for ln in (card_text or "").splitlines() if ln.strip()]
                name = clean_tournament_name(lines[0]) if lines else clean_tournament_name(href)

            # Date text shown on the card (e.g. "15/02/2026 to 21/02/2026")
            date_str = ""
            m = re.search(r"\b\d{2}/\d{2}/\d{4}\b(?:\s*(?:to|-)\s*\d{2}/\d{2}/\d{4}\b)?", card_text)
            if m:
                date_str = m.group(0).strip()

            tournaments.append({
                "name": name,
                "date": date_str,     # later parsing already uses "first date" logic
                "event": "",          # filled later via "My Entry"
                "link": href,
                "status": None,
                "grade": None,
                "draw_size": None,
                "entries": []
            })

            print(f"  ✓ Found: {name} | {date_str if date_str else '(no card date)'}")

    except Exception as e:
        print(f"Error in scrape_my_entries(/tournaments): {e}")
        try:
            page.screenshot(path="debug_scrape_my_tournaments.png", full_page=True)
            with open("debug_scrape_my_tournaments.html", "w", encoding="utf-8") as f:
                f.write(page.content())
        except Exception:
            pass

    return tournaments



def enrich_tournament_with_my_entry_event(page, tournament):
    """
    Open tournament page, click 'My Entry', and extract ALL entered event codes
    (e.g. ['9U BS', '10U BS'] when entered in multiple events).

    Stores the list in tournament["events"] (plural).
    Also fills tournament['date'] from Fact Sheet start date if it is missing.
    """
    try:
        page.goto(tournament["link"], wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(1200)

        # Click "My Entry" (seen as a nav-link style button on the tournament page)
        my_entry = page.get_by_role("link", name=re.compile(r"^\s*My Entry\s*$", re.I))
        if my_entry.count() == 0:
            my_entry = page.get_by_role("button", name=re.compile(r"^\s*My Entry\s*$", re.I))
        my_entry.first.click(timeout=15000)
        page.wait_for_load_state("domcontentloaded", timeout=20000)
        page.wait_for_timeout(800)

        # On the My Entry view, find the "Events" dt and read the sibling dd
        dt_events = page.locator("dt.list__label", has_text=re.compile(r"^\s*Events\s*$", re.I)).first
        if dt_events.count() == 0:
            # fallback for loud label variant
            dt_events = page.locator("dt", has_text=re.compile(r"^\s*Events\s*$", re.I)).first

        event_text = ""
        if dt_events.count() > 0:
            dd = dt_events.locator("xpath=following-sibling::dd[1]").first
            event_text = (dd.inner_text() or "").strip()

        # Extract ALL event codes like '9U BS', '10U BS', '14U GS' etc.
        event_codes = re.findall(r"\b\d{1,2}U\s+[A-Z]{1,3}\b", event_text)
        event_codes = [e.strip() for e in event_codes]

        if event_codes:
            tournament["events"] = event_codes
            print(f"  ✓ Entered event(s): {event_codes}")
        else:
            print("  ⚠️  Could not extract entered event(s) from 'My Entry' view.")
            tournament["events"] = []

        # If date is missing, grab it from Fact Sheet (Start date)
        if not tournament.get("date"):
            try:
                fact_sheet_tab = page.get_by_role("link", name=re.compile(r"^\s*Fact Sheet\s*$", re.I))
                if fact_sheet_tab.count() == 0:
                    fact_sheet_tab = page.locator('a:has-text("Fact Sheet")')
                fact_sheet_tab.first.click(timeout=15000)
                page.wait_for_load_state("domcontentloaded", timeout=20000)
                page.wait_for_timeout(800)

                dt_start = page.locator("dt.list__label", has_text=re.compile(r"^\s*Start date\s*$", re.I)).first
                if dt_start.count() > 0:
                    dd_start = dt_start.locator("xpath=following-sibling::dd[1]").first
                    start_txt = (dd_start.inner_text() or "").strip()
                    # Expect dd/mm/yyyy
                    m2 = re.search(r"\b\d{2}/\d{2}/\d{4}\b", start_txt)
                    if m2:
                        tournament["date"] = m2.group(0)
                        print(f"  ✓ Date (Start date): {tournament['date']}")
            except Exception:
                pass

    except Exception as e:
        print(f"  ⚠️  Could not enrich tournament with 'My Entry' info: {e}")


def scrape_tournament_details(page, tournament, scrape_entries=True):
    """Scrape tournament details from Fact Sheet; optionally scrape entries."""
    print(f"\n{'='*60}")
    print(f"Scraping details for: {tournament['name']}")
    print(f"Link: {tournament['link']}")
    print(f"{'='*60}")

    try:
        page.goto(tournament["link"], wait_until="networkidle", timeout=30000)
        page.wait_for_timeout(1200)

        # Fact Sheet tab
        print("  Navigating to Fact Sheet...")
        try:
            fact_sheet_tab = page.query_selector('a:has-text("Fact Sheet")') \
                             or page.query_selector('tab:has-text("Fact Sheet")') \
                             or page.query_selector('a[href*="Factsheet"]')

            if fact_sheet_tab:
                fact_sheet_tab.click()
                page.wait_for_load_state("networkidle")
                page.wait_for_timeout(800)
            else:
                print("  WARNING: Could not find Fact Sheet tab")
        except Exception as e:
            print(f"  WARNING: Could not click Fact Sheet tab: {e}")

        # Event Information section
        print(f"  Looking for event: {tournament['event']}")
        page.wait_for_selector("text=/Event Information/i", timeout=10000)
        page.wait_for_timeout(600)

        # Find event "button" by exact match (per your preference)
        all_buttons = page.query_selector_all("button")
        event_button = None
        for btn in all_buttons:
            try:
                if btn.inner_text().strip() == tournament["event"]:
                    event_button = btn
                    break
            except Exception:
                continue

        if event_button:
            print(f"  Found event button: {tournament['event']}")
            event_button.click()
            page.wait_for_timeout(1500)

            page_html = page.content()


            # Grade (scoped to the selected event)
            try:
                raw = extract_fact_sheet_value(page, tournament["event"], "Grade")
                if raw:
                    m = re.search(r"Grade\s*\d+", raw, re.I)
                    tournament["grade"] = m.group(0) if m else raw
                else:
                    tournament["grade"] = None
                print(f"  ✓ Grade: {tournament['grade']}" if tournament["grade"] else "  ✗ Grade not found")
            except Exception as e:
                print(f"  ✗ Error extracting grade: {e}")
                tournament["grade"] = None

            # Proposed Draw Size (scoped to the selected event)
            # Defaults to 16 if not available or value contains no numeric data
            try:
                raw = extract_fact_sheet_value(page, tournament["event"], "Proposed Draw Size")
                if raw:
                    nums = re.findall(r"\d+", raw)
                    tournament["draw_size"] = str(max(int(n) for n in nums)) if nums else 16
                else:
                    tournament["draw_size"] = 16
                if tournament["draw_size"] == 16:
                    print(f"  ✗ Draw Size not found or non-numeric — defaulting to 16")
                else:
                    print(f"  ✓ Draw Size: {tournament['draw_size']}")
            except Exception as e:
                print(f"  ✗ Error extracting draw size: {e} — defaulting to 16")
                tournament["draw_size"] = 16

            # Venue (tournament-level field; first line only to drop trailing "Open in Maps" etc.)
            try:
                raw_venue = extract_fact_sheet_value(page, tournament["event"], "Venue")
                tournament["venue"] = (raw_venue or "").split("\n")[0].strip()
                print(f"  ✓ Venue: {tournament['venue']}" if tournament["venue"] else "  ✗ Venue not found")
            except Exception as e:
                print(f"  ✗ Error extracting venue: {e}")
                tournament["venue"] = ""

            # Event date from Timings (scoped to the selected event)
            # Uses Start date + Timings to calculate the actual event dates
            # Falls back to event-scoped Start date + End date if Timings unavailable
            try:
                start_date_raw = extract_fact_sheet_value(page, tournament["event"], "Start date")
                timings_raw = extract_fact_sheet_value(page, tournament["event"], "Timings")
                event_date = None
                if timings_raw:
                    event_date = parse_timings_to_event_date(timings_raw, start_date_raw or tournament.get("date", ""))
                    if event_date:
                        print(f"  ✓ Event date (from Timings): {event_date}  [Timings: {timings_raw}]")
                    else:
                        print(f"  ℹ️  Timings field present but could not parse dates: {timings_raw}")

                # Fallback: use event-scoped Start date + End date from Fact Sheet
                if not event_date and start_date_raw:
                    start_m = re.search(r"\b\d{2}/\d{2}/\d{4}\b", start_date_raw)
                    if start_m:
                        end_date_raw = extract_fact_sheet_value(page, tournament["event"], "End date")
                        end_m = re.search(r"\b\d{2}/\d{2}/\d{4}\b", end_date_raw or "")
                        if end_m and end_m.group(0) != start_m.group(0):
                            event_date = f"{start_m.group(0)} - {end_m.group(0)}"
                        else:
                            event_date = start_m.group(0)
                        print(f"  ✓ Event date (from Start/End date): {event_date}")

                if event_date:
                    # Extract start time if available in Timings (stored separately)
                    start_time = extract_start_time(timings_raw)
                    if start_time:
                        tournament["start_time"] = start_time
                        print(f"  ✓ Start time: {start_time}")
                    tournament["date"] = event_date
                elif not timings_raw and not start_date_raw:
                    print(f"  ℹ️  No Timings or Start date found — keeping existing date: {tournament.get('date', '')}")
            except Exception as e:
                print(f"  ⚠️  Error extracting event date from Timings: {e}")

            # Closing deadline -> Status (scoped to the selected event)
            try:
                closing_deadline = extract_fact_sheet_value(page, tournament["event"], "Closing deadline")
                tournament["status"] = determine_status(closing_deadline, tournament["date"])
                print(f"  ✓ Status: {tournament['status']}")

                # Store closing date string (same format as watchlist: "dd/mm/yyyy HH:MM")
                closing_dt = parse_closing_deadline(closing_deadline)
                if closing_dt:
                    tournament["closing_date"] = closing_dt.strftime("%d/%m/%Y %H:%M")
                    print(f"  ✓ Closing date: {tournament['closing_date']}")
            except Exception as e:
                print(f"  ✗ Error determining status: {e}")
                tournament["status"] = "Unknown"



        else:
            print(f"  WARNING: Could not find event button for {tournament['event']}")

    except Exception as e:
        print(f"  ERROR in scrape_tournament_details: {e}")

    if scrape_entries:
        scrape_tournament_entries(page, tournament)


def scrape_tournament_entries(page, tournament):
    """
    Scrape entries from Events tab.

    Rules:
    - Go to Events tab
    - Click the specific event (e.g. '14U BS')
    - Ignore Withdrawn entirely
    - If 'Entries (N)' exists -> ONLY use that table:
        - Use Player column only (ignore Maindraw column)
        - Entry Date blank
    - Else use 'Online entries (N)' table:
        - Use Name + Date of entry
    """
    print(f"\n  {'---'*20}")
    print(f"  Extracting entries for: {tournament['name']}")

    try:
        # Navigate to Events tab
        print("  Navigating to Events tab...")
        events_tab = page.locator("a", has_text=re.compile(r'^\s*Events\s*$', re.I)).first
        if events_tab.count() == 0:
            events_tab = page.locator('a[href*="/events"]').first
        if events_tab.count() == 0:
            print("  WARNING: Could not find Events tab")
            tournament["entries"] = []
            return

        events_tab.click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(800)

        # Click the specific event link (e.g., 14U BS)
        event_text = (tournament.get("event") or "").strip()
        print(f"  Looking for event link: {event_text}")

        # Try exact-ish match first, then fallback
        event_link = page.locator("a", has_text=re.compile(rf"^\s*{re.escape(event_text)}\s*$", re.I)).first
        if event_link.count() == 0:
            event_link = page.locator(f'a:has-text("{event_text}")').first

        if event_link.count() == 0:
            print(f"  WARNING: Could not find event link for {event_text}")
            tournament["entries"] = []
            return

        event_link.click()
        page.wait_for_load_state("domcontentloaded")
        page.wait_for_timeout(1200)
        tournament["link"] = page.url  # replace overview URL with direct entries URL

        # ---------- Prefer Entries table ----------
        entries_table = page.locator("table.ruler").filter(
            has=page.locator("caption", has_text=re.compile(r"^\s*Entries\s*\(\d+\)\s*$", re.I))
        ).first

        if entries_table.count() > 0:
            print("  ✓ Found 'Entries' table (using Player column only)")

            rows = entries_table.locator("tbody tr")
            n = rows.count()
            entries = []
            seen = set()

            # PASS 1 (existing logic): assume Player is in td[1] (works for Maindraw-style)
            for i in range(n):
                tr = rows.nth(i)
                tds = tr.locator("td")
                if tds.count() < 2:
                    continue

            # gate rows based on the first column label
                raw_label = tds.nth(0).inner_text() or ""
                label = raw_label.replace("\u00a0", " ").strip().lower()

            # Ignore reserves entirely
                if label.startswith("reserve"):
                    continue

                if not (label.startswith("maindraw") or label.startswith("qualification")):
                    continue


                # td[1] = Player
                player_td = tds.nth(1)
                player_link = player_td.locator("a").first

                if player_link.count() > 0:
                    player_name = player_link.inner_text().strip()
                else:
                    player_name = player_td.inner_text().strip()

                if player_name and player_name not in seen:
                    seen.add(player_name)
                    entries.append({"name": player_name, "entry_date": ""})

            # PASS 2 (exception): only if PASS 1 returns nothing.
            # Accept Qualification* in td[0], ignore Reserve* in td[0]
            if not entries:
                for i in range(n):
                    tr = rows.nth(i)
                    tds = tr.locator("td")
                    if tds.count() < 2:
                        continue

                    col0 = tds.nth(0).inner_text().strip()

                    # Ignore reserve rows
                    if re.search(r"^\s*reserve\b", col0, re.I):
                        continue

                    # Only take qualification rows
                    if not re.search(r"^\s*qualification\b", col0, re.I):
                        continue

                    player_td = tds.nth(1)
                    player_link = player_td.locator("a").first

                    if player_link.count() > 0:
                        player_name = player_link.inner_text().strip()
                    else:
                        player_name = player_td.inner_text().strip()

                    if player_name and player_name not in seen:
                        seen.add(player_name)
                        entries.append({"name": player_name, "entry_date": ""})


            tournament["entries"] = entries
            print(f"  ✓ Extracted {len(entries)} players from 'Entries'")

            # Override draw size if actual entry count differs from Fact Sheet value
            if entries:
                try:
                    current_draw = int(tournament.get("draw_size", 0))
                except (ValueError, TypeError):
                    current_draw = 0
                if len(entries) != current_draw:
                    print(f"  ⚠️  Draw size override: {current_draw} → {len(entries)} (from entry count)")
                    tournament["draw_size"] = str(len(entries))

            print(f"  {'---'*20}")
            return

        # ---------- Otherwise use Online entries table ----------
        online_table = page.locator("table.ruler").filter(
            has=page.locator("caption", has_text=re.compile(r"^\s*Online entries\s*\(\d+\)\s*$", re.I))
        ).first

        if online_table.count() == 0:
            print("  ✗ No 'Entries' or 'Online entries' table found")
            tournament["entries"] = []
            return

        print("  ✓ Found 'Online entries' table (Name + Date of entry)")

        rows = online_table.locator("tbody tr")
        n = rows.count()
        entries = []

        for i in range(n):
            tr = rows.nth(i)
            tds = tr.locator("td")
            if tds.count() < 2:
                continue

            name = tds.nth(0).inner_text().strip()
            entry_date = tds.nth(1).inner_text().strip()

            if name and name.lower() not in ("name", "player"):
                entries.append({"name": name, "entry_date": entry_date})

        tournament["entries"] = entries
        print(f"  ✓ Extracted {len(entries)} entries from 'Online entries'")
        print(f"  {'---'*20}")

    except Exception as e:
        print(f"  ✗ ERROR in scrape_tournament_entries: {e}")
        tournament["entries"] = []



def update_tournament_entries_only(page, tournament):
    """For existing tournaments, navigate to the tournament then update the entries list."""
    print(f"\nUpdating entries for existing tournament: {tournament['name']}")

    try:
        page.goto(tournament["link"], wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(800)
    except Exception as e:
        print(f"  WARNING: Could not navigate to tournament page for entries scrape: {e}")

    scrape_tournament_entries(page, tournament)



# ---------------------------
# Main
# ---------------------------

def main():
    print("=== LTA Tournament Scraper ===\n")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        # Navigate to home page
        print("Navigating to LTA competitions site...")
        page.goto("https://competitions.lta.org.uk/", wait_until="domcontentloaded")
        page.wait_for_timeout(800)

        # Dismiss cookies
        print("Checking for cookie banner...")
        dismiss_cookie_banner(page)

        # Login (automated if env vars exist; else manual)
        login_if_needed(page)

        # Ensure we're at a stable state before scraping
        try:
            page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass

                # Scrape current entries list
        tournaments = scrape_my_entries(page)
        print(f"\nFound {len(tournaments)} tournaments in 'My tournaments'")

        # Process each tournament (FULL scrape every run; no existing-tournament checks)
        # If a tournament has multiple entered events, expand it into one dict per event.
        expanded_tournaments = []
        for tournament in tournaments:
            print(f"\nEnriching tournament with entered event(s) via 'My Entry': {tournament.get('name','')}")
            enrich_tournament_with_my_entry_event(page, tournament)

            events = tournament.get("events") or []
            if len(events) <= 1:
                # Single event (or none found) — keep as-is, set event field for downstream use
                tournament["event"] = events[0] if events else ""
                expanded_tournaments.append(tournament)
            else:
                # Multiple events — create a separate tournament dict per event
                print(f"  ↳ Multiple events found ({events}) — expanding into {len(events)} entries")
                import copy
                for event_code in events:
                    t_copy = copy.deepcopy(tournament)
                    t_copy["event"] = event_code
                    t_copy["events"] = [event_code]
                    expanded_tournaments.append(t_copy)

        tournaments = expanded_tournaments

        for tournament in tournaments:
            # Full scrape (Fact Sheet + Entries)
            scrape_tournament_details(page, tournament, scrape_entries=True)

# Filter out "Played"
        active_tournaments = [t for t in tournaments if t.get("status") != "Played"]
        print(f"\n{len(active_tournaments)} active tournaments (Open/Closed)")
        print(f"{len(tournaments) - len(active_tournaments)} tournaments filtered out (Played)")

        # Excel: clear existing data and write fresh blocks
#        write_fresh_to_excel(
#            active_tournaments,
#            filename=OUTPUT_XLSX,
#            sheet_name=OUTPUT_SHEET
#        )


        logger = setup_logger()

        write_wide_to_sheet_via_webapp(
            tournaments=active_tournaments,
            sheet_tab="DYLAN_TOURNAMENTS",
            clear_first=True,
            logger=logger,
        )




        print("\n✅ Scraping complete!")
        browser.close()


if __name__ == "__main__":
    main()